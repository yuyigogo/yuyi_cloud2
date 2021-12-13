import os
import re
from abc import ABC, abstractmethod
from collections import defaultdict
from io import BufferedReader, BytesIO, TextIOBase
from logging import getLogger
from typing import Dict, List, Optional, Union

import xlrd
from bson import ObjectId
from django.core.files.uploadedfile import UploadedFile
from openpyxl import Workbook as OpenWorkbook
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE, Cell
from openpyxl.reader.excel import SUPPORTED_FORMATS
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError, InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from common.error_code import StatusCode
from common.framework.exception import APIException

logger = getLogger(__name__)


class Workbook(ABC):
    def __init__(self, _workbook):
        self._workbook = _workbook

    @abstractmethod
    def get_columns(self, sheet_name=None):
        pass

    @classmethod
    def open(cls, excel_file_name=None, excel_file=None):
        # try openpyxl parse excel file, and old xls file will raise InvalidFileException, parse use xlrd
        try:
            return OpenpyxlWorkbook.open(excel_file_name, excel_file)
        except InvalidFileException:
            return XlrdWorkbook.open(excel_file_name, excel_file)


class XlrdWorkbook(Workbook):
    def get_columns(self, sheet_name=None):
        if sheet_name is None:
            sheet = self._workbook.sheet_by_index(0)
        else:
            sheet = self._workbook.sheet_by_name(sheet_name)

        for r_i in range(sheet.nrows):
            result = []
            for c_i in range(sheet.ncols):
                cell = sheet.cell(r_i, c_i)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate_as_datetime(value, self._workbook.datemode)
                elif isinstance(value, (float, int)):
                    xf = self._workbook.xf_list[cell.xf_index]
                    fmt = self._workbook.format_map[xf.format_key]
                    number_format = fmt.format_str
                    value = real_number_str(value, number_format)
                result.append(value)
            yield result

    @classmethod
    def open(cls, excel_file_name=None, excel_file=None):
        if excel_file is not None:
            if isinstance(excel_file, (BufferedReader, UploadedFile)):
                file_contents = excel_file.read()
            elif isinstance(excel_file, bytes):
                file_contents = excel_file
            elif isinstance(excel_file, TextIOBase):
                raise ValueError("Open excel file with binary mode")
            else:
                raise ValueError("Unknown excel_file type")
            workbook = xlrd.open_workbook(
                file_contents=file_contents, formatting_info=True
            )
        elif excel_file_name is not None:
            workbook = xlrd.open_workbook(
                filename=excel_file_name, formatting_info=True
            )
        else:
            raise ValueError("import excel must need excel_file_name or excel_file")
        return XlrdWorkbook(workbook)


class OpenpyxlWorkbook(Workbook):
    def get_columns(self, sheet_name=None):
        if sheet_name is None:
            sheet = list(self._workbook)[0]
        else:
            sheet = self._workbook[sheet_name]
        max_column = sheet.max_column
        for r_i in range(sheet.max_row):
            result = []
            for c_i in range(max_column):
                cell = sheet.cell(r_i + 1, c_i + 1)
                if isinstance(cell, MergedCell):
                    raise APIException(
                        "import excel contains merged cells",
                        code=StatusCode.EXCEL_CONTAINS_MERGED_CELL.value,
                    )
                value = cell.value
                if isinstance(value, (float, int)):
                    if cell.number_format.endswith("%"):
                        value *= 100
                    value = real_number_str(value, cell.number_format)
                if value is None:
                    value = ""
                result.append(value)
            yield result

    @classmethod
    def open(cls, excel_file_name=None, excel_file=None):
        if excel_file is not None:
            if hasattr(excel_file, "name"):
                file_format = os.path.splitext(excel_file.name)[-1].lower()
                if file_format not in SUPPORTED_FORMATS:
                    raise InvalidFileException()
            workbook = load_workbook(excel_file)
        elif excel_file_name is not None:
            workbook = load_workbook(excel_file_name)
        else:
            raise ValueError("Invalid excel file")
        return cls(workbook)


re_maybe_numfmt = re.compile("[0#.,]*[0#][0#.,]*")


def extract_number_format(s_fmt):
    """
    from numfmt.py in http://uucode.com/blog/2013/10/22/using-xlrd-and-formatting-excel-numbers/
    """
    # If don't know what does the format "Standard/GENERAL" mean.
    # As far as I understand, the presentation can differ depending
    # on the locale and user settings. Here is a my proposal.
    if "GENERAL" == s_fmt.upper():
        return None, "#", "#"
    # Find the number-part
    m = re_maybe_numfmt.search(s_fmt)
    if m is None:
        return None, "#", "#"
    s_numfmt = str(m.group(0))
    # Only one comma
    pos_comma = s_numfmt.find(",")
    if pos_comma > -1:
        pos2 = s_numfmt.find(",", pos_comma + 1)
        if pos2 > -1:
            return None, "#", "#"
    # Only one dot
    pos_dot = s_numfmt.find(".")
    if pos_dot > -1:
        pos2 = s_numfmt.find(".", pos_dot + 1)
        if pos2 > -1:
            return None, "#", "#"
    # Exactly three positions between comma and dot
    if pos_comma > -1:
        pos2 = pos_dot if pos_dot > -1 else len(s_numfmt)
        if pos2 - pos_comma != 4:
            return None, "#", "#"
    # Create parts
    (part_above1000, part_below1000, part_below1) = (None, None, None)
    if pos_dot > -1:
        part_below1 = s_numfmt[pos_dot + 1 :]
        s_numfmt = s_numfmt[:pos_dot]
    if pos_comma > -1:
        part_above1000 = s_numfmt[:pos_comma]
        part_below1000 = s_numfmt[pos_comma + 1 :]
    else:
        part_below1000 = s_numfmt
    return part_above1000, part_below1000, part_below1


def real_number_str(f, number_format):
    (part_above1000, part_below1000, part_below1) = extract_number_format(number_format)
    is_negative = f < 0
    f = abs(f)
    # Float to string with a minimal precision after comma.
    # Filling the integer part with '0' at left doesn't work for %f.
    precision = len(part_below1) if part_below1 else 0
    if part_below1 == "#":
        s_f = str(float(f))
    else:
        s_fmt = "%." + str(precision) + "f"
        s_f = s_fmt % f
    # Postprocessing. Drop insignificant zeros.
    while precision:
        if "0" == part_below1[precision - 1]:
            break
        if "0" != s_f[-1]:
            break
        s_f = s_f[:-1]
        precision = precision - 1
    if "." == s_f[-1]:
        s_f = s_f[:-1]
    if is_negative:
        s_f = "-" + s_f
    return s_f


HEADER_BLUE = "header_blue"
LIGHT_GREEN = "light_green"
LIGHT_YELLOW = "light_yellow"
PATTERN_FILLS = {
    HEADER_BLUE: PatternFill("solid", fgColor="538ed5"),
    LIGHT_GREEN: PatternFill("solid", fgColor="c6efce"),
    LIGHT_YELLOW: PatternFill("solid", fgColor="ffeb9c"),
}
GREEN_FONT = "green_font"
YELLOW_FONT = "yellow_font"
TEXT_FONTS = {
    GREEN_FONT: Font(color="006100"),
    YELLOW_FONT: Font(color="9c5600"),
}


def set_cell_value(cell: Cell, value: Union[None, str, int, ObjectId]) -> None:
    try:
        cell.value = value
    except IllegalCharacterError as e:
        logger.info(f"IllegalCharacterError: {e=}")
        cell.value = ILLEGAL_CHARACTERS_RE.sub(r"", value)


def fill_xlsx_data(
    ws: Worksheet,
    data: List[List[Union[str, dict]]],
    color_lines: Optional[list] = None,
    bold_lines: Optional[list] = None,
) -> None:
    """
    :param ws:
    :param data: [[title1, title2], [val1, val2], ]
    :param color_lines:
    :param bold_lines
    :return:
    """
    color_lines = color_lines or []
    bold_lines = bold_lines or []
    column_widths = defaultdict(int)
    for i, line in enumerate(data, 1):
        for j, v in enumerate(line, 1):
            cell = ws.cell(i, j)
            if isinstance(v, dict):
                set_cell_value(cell, v["value"])
                cell.fill = PATTERN_FILLS[v["fill"]]
                cell.font = TEXT_FONTS[v["font"]]
            else:
                set_cell_value(cell, v)
            # color line
            if i in color_lines:
                cell.fill = PATTERN_FILLS[HEADER_BLUE]
                cell.font = Font(bold=True, color="FFFFFF")
            if i in bold_lines:
                cell.font = Font(bold=True)
            # collect max column width, add two space for display
            if v:
                column_widths[j] = max(len(v) + 2, column_widths[j])

    # adjust column width
    for j, width in column_widths.items():
        ws.column_dimensions[get_column_letter(j)].width = width


def data_to_xlsx(
    data: List[List[Union[str, dict]]],
    color_lines: Optional[list] = None,
    bold_lines: Optional[list] = None,
) -> bytes:
    color_lines = color_lines or []
    bold_lines = bold_lines or []
    wb = OpenWorkbook()
    ws = wb.active
    fill_xlsx_data(ws, data, color_lines, bold_lines)
    with BytesIO() as output:
        wb.save(output)
        return output.getvalue()


def sheet_data_to_xlsx(
    sheet_data: List[Dict[str, Union[str, List[List[Union[str, dict]]]]]],
    color_lines: list = None,
    bold_lines: list = None,
) -> bytes:
    """
    :param sheet_data:
    [{
        "sheet_title": "sheet_title1",
        "sheet_data": [
            [title1, title2],
            [val1, val2],
        ],
        "color_lines": [1, 2],
        "bold_lines": [3, 4]
    },
    {
        "sheet_title": "sheet_title2",
        "sheet_data": [
            [title1, title2],
            [val1, val2],
        ]
    }
    ]
    :param color_lines: [1]
    :param bold_lines: [1]
    :return:
    """
    color_lines = color_lines or []
    bold_lines = bold_lines or []
    wb = OpenWorkbook()
    wb._sheets = []
    for sd in sheet_data:
        ws = wb.create_sheet(sd["sheet_title"])
        ws_color_lines = sd.get("color_lines", color_lines)
        ws_bold_lines = sd.get("bold_lines", bold_lines)
        fill_xlsx_data(
            ws, sd["sheet_data"], color_lines=ws_color_lines, bold_lines=ws_bold_lines
        )
    with BytesIO() as output:
        wb.save(output)
        return output.getvalue()
